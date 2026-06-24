"""Data ingestion layer for Connected Mode.

Architecture (see docs/INTEGRATIONS.md):

    External Systems -> Connector Layer -> Validation Layer ->
    Transformation Layer -> Database -> Risk Engine / Analytics / AI Copilot

This module implements:
  * Integration template catalog + connector seeding / sync simulation
  * A field-spec registry describing how each importable entity maps + validates
  * CSV / Excel parsing
  * Validation + transformation + commit into the existing domain tables
  * Operating-mode (Demo vs Connected) persistence
"""
from __future__ import annotations

import csv
import io
import random
from datetime import datetime, timedelta, timezone

from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.models import (
    AppSetting,
    DataSource,
    ImportJob,
    Inventory,
    Product,
    Shipment,
    Supplier,
    Warehouse,
)
from app.models.enums import (
    ConnectorHealth,
    ConnectorStatus,
    ConnectorType,
    ImportStatus,
    OperatingMode,
    ShipmentStatus,
)


def _utcnow() -> datetime:
    return datetime.now(timezone.utc)


# --------------------------------------------------------------------------- #
# Integration template catalog
# --------------------------------------------------------------------------- #
INTEGRATION_TEMPLATES: list[dict] = [
    {"type": ConnectorType.SAP_ERP.value, "name": "SAP ERP", "category": "ERP",
     "description": "Sync purchase orders, materials and supplier master data from SAP S/4HANA.",
     "auth_methods": ["OAuth2", "API Key", "Basic"]},
    {"type": ConnectorType.ORACLE_ERP.value, "name": "Oracle ERP Cloud", "category": "ERP",
     "description": "Ingest procurement, inventory and supplier records from Oracle Fusion.",
     "auth_methods": ["OAuth2", "API Key"]},
    {"type": ConnectorType.SALESFORCE_CRM.value, "name": "Salesforce CRM", "category": "CRM",
     "description": "Pull demand signals, accounts and order pipeline from Salesforce.",
     "auth_methods": ["OAuth2", "API Key"]},
    {"type": ConnectorType.MS_DYNAMICS.value, "name": "Microsoft Dynamics 365", "category": "ERP/CRM",
     "description": "Connect Dynamics 365 Supply Chain for orders and inventory positions.",
     "auth_methods": ["OAuth2", "API Key"]},
    {"type": ConnectorType.WMS.value, "name": "Warehouse Management System", "category": "Warehouse",
     "description": "Stream real-time stock levels, bin locations and receipts from your WMS.",
     "auth_methods": ["API Key", "Basic"]},
    {"type": ConnectorType.TMS.value, "name": "Transportation Management System", "category": "Transportation",
     "description": "Ingest shipment tracking, carrier events and ETAs from your TMS.",
     "auth_methods": ["API Key", "OAuth2"]},
    {"type": ConnectorType.REST_API.value, "name": "Generic REST API", "category": "Custom",
     "description": "Connect any RESTful endpoint with a configurable schema mapping.",
     "auth_methods": ["API Key", "Bearer Token", "Basic", "None"]},
]

_TEMPLATE_BY_TYPE = {t["type"]: t for t in INTEGRATION_TEMPLATES}


# --------------------------------------------------------------------------- #
# Importable entity field specs
# --------------------------------------------------------------------------- #
# type: one of str | int | float | datetime | enum:<EnumName>
ENTITY_SPECS: dict[str, dict] = {
    "suppliers": {
        "label": "Suppliers",
        "fields": [
            {"name": "name", "label": "Name", "required": True, "type": "str"},
            {"name": "country", "label": "Country", "required": True, "type": "str"},
            {"name": "region", "label": "Region", "required": True, "type": "str"},
            {"name": "category", "label": "Category", "required": False, "type": "str", "default": "General"},
            {"name": "supplier_score", "label": "Supplier Score", "required": False, "type": "float", "default": 80.0},
            {"name": "delivery_reliability", "label": "Delivery Reliability", "required": False, "type": "float", "default": 90.0},
            {"name": "average_delay_days", "label": "Avg Delay (days)", "required": False, "type": "float", "default": 1.0},
            {"name": "order_fulfillment_rate", "label": "Fulfillment Rate", "required": False, "type": "float", "default": 95.0},
            {"name": "defect_rate", "label": "Defect Rate", "required": False, "type": "float", "default": 1.0},
        ],
    },
    "warehouses": {
        "label": "Warehouses",
        "fields": [
            {"name": "name", "label": "Name", "required": True, "type": "str"},
            {"name": "location", "label": "Location", "required": True, "type": "str"},
            {"name": "region", "label": "Region", "required": True, "type": "str"},
            {"name": "latitude", "label": "Latitude", "required": False, "type": "float", "default": 0.0},
            {"name": "longitude", "label": "Longitude", "required": False, "type": "float", "default": 0.0},
            {"name": "capacity", "label": "Capacity", "required": True, "type": "int"},
            {"name": "current_inventory", "label": "Current Inventory", "required": False, "type": "int", "default": 0},
        ],
    },
    "products": {
        "label": "Products",
        "fields": [
            {"name": "sku", "label": "SKU", "required": True, "type": "str", "unique": True},
            {"name": "name", "label": "Name", "required": True, "type": "str"},
            {"name": "category", "label": "Category", "required": False, "type": "str", "default": "General"},
            {"name": "unit_cost", "label": "Unit Cost", "required": True, "type": "float"},
            {"name": "unit_price", "label": "Unit Price", "required": True, "type": "float"},
            {"name": "lead_time_days", "label": "Lead Time (days)", "required": False, "type": "int", "default": 14},
        ],
    },
    "shipments": {
        "label": "Shipments",
        "fields": [
            {"name": "reference", "label": "Reference", "required": True, "type": "str", "unique": True},
            {"name": "origin", "label": "Origin", "required": True, "type": "str"},
            {"name": "destination", "label": "Destination", "required": True, "type": "str"},
            {"name": "carrier", "label": "Carrier", "required": True, "type": "str"},
            {"name": "status", "label": "Status", "required": False, "type": "enum:ShipmentStatus", "default": "in_transit"},
            {"name": "units", "label": "Units", "required": False, "type": "int", "default": 0},
            {"name": "value_usd", "label": "Value (USD)", "required": False, "type": "float", "default": 0.0},
        ],
    },
    "inventory": {
        "label": "Inventory",
        "fields": [
            {"name": "warehouse_name", "label": "Warehouse Name", "required": True, "type": "str", "fk": "warehouse"},
            {"name": "product_sku", "label": "Product SKU", "required": True, "type": "str", "fk": "product"},
            {"name": "quantity", "label": "Quantity", "required": True, "type": "int"},
            {"name": "reorder_point", "label": "Reorder Point", "required": False, "type": "int", "default": 0},
            {"name": "safety_stock", "label": "Safety Stock", "required": False, "type": "int", "default": 0},
            {"name": "max_stock", "label": "Max Stock", "required": False, "type": "int", "default": 0},
            {"name": "avg_daily_demand", "label": "Avg Daily Demand", "required": False, "type": "float", "default": 0.0},
        ],
    },
}


def entity_spec(entity: str) -> dict:
    if entity not in ENTITY_SPECS:
        raise ValueError(f"Unknown entity '{entity}'")
    return ENTITY_SPECS[entity]


# --------------------------------------------------------------------------- #
# Parsing
# --------------------------------------------------------------------------- #
def parse_upload(filename: str, content: bytes, sheet: str | None = None) -> dict:
    """Return {columns, rows, sheets} for a CSV or XLSX upload."""
    lower = (filename or "").lower()
    if lower.endswith(".xlsx") or lower.endswith(".xlsm"):
        return _parse_excel(content, sheet)
    return _parse_csv(content)


def _parse_csv(content: bytes) -> dict:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    columns = list(reader.fieldnames or [])
    rows = [dict(r) for r in reader]
    return {"columns": columns, "rows": rows, "sheets": []}


def _parse_excel(content: bytes, sheet: str | None) -> dict:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("openpyxl is required for Excel import") from exc

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheets = wb.sheetnames
    ws = wb[sheet] if sheet and sheet in sheets else wb[sheets[0]]

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return {"columns": [], "rows": [], "sheets": sheets}
    columns = [str(c) if c is not None else f"column_{i}" for i, c in enumerate(header)]
    rows: list[dict] = []
    for raw in rows_iter:
        if raw is None or all(c is None for c in raw):
            continue
        row = {columns[i]: ("" if i >= len(raw) or raw[i] is None else raw[i]) for i in range(len(columns))}
        rows.append(row)
    wb.close()
    return {"columns": columns, "rows": rows, "sheets": sheets}


# --------------------------------------------------------------------------- #
# Mapping suggestion
# --------------------------------------------------------------------------- #
def _norm(s: str) -> str:
    return "".join(ch for ch in str(s).lower() if ch.isalnum())


def suggest_mapping(entity: str, columns: list[str]) -> dict[str, str | None]:
    spec = entity_spec(entity)
    norm_cols = {_norm(c): c for c in columns}
    mapping: dict[str, str | None] = {}
    for field in spec["fields"]:
        fn = field["name"]
        candidates = {_norm(fn), _norm(field["label"])}
        # a few common aliases
        candidates |= {_norm(fn.replace("_", " "))}
        match = next((norm_cols[c] for c in candidates if c in norm_cols), None)
        mapping[fn] = match
    return mapping


# --------------------------------------------------------------------------- #
# Validation + transformation
# --------------------------------------------------------------------------- #
def _coerce(value, ftype: str):
    if value is None:
        raise ValueError("missing")
    s = str(value).strip()
    if s == "":
        raise ValueError("empty")
    if ftype == "str":
        return s
    if ftype == "int":
        return int(float(s))
    if ftype == "float":
        return float(s)
    if ftype == "datetime":
        return datetime.fromisoformat(s)
    if ftype.startswith("enum:"):
        return s.lower()
    return s


def validate_rows(entity: str, rows: list[dict], mapping: dict[str, str | None]) -> dict:
    """Validate (without writing). Returns sample of normalized rows + errors."""
    spec = entity_spec(entity)
    errors: list[dict] = []
    valid = 0
    normalized_sample: list[dict] = []

    for idx, raw in enumerate(rows):
        norm, row_errors = _transform_row(spec, raw, mapping)
        if row_errors:
            errors.append({"row": idx + 1, "errors": row_errors})
        else:
            valid += 1
            if len(normalized_sample) < 10:
                normalized_sample.append(norm)

    return {
        "total": len(rows),
        "valid": valid,
        "rejected": len(rows) - valid,
        "errors": errors[:50],
        "sample": normalized_sample,
    }


def _transform_row(spec: dict, raw: dict, mapping: dict[str, str | None]) -> tuple[dict, list[str]]:
    out: dict = {}
    errs: list[str] = []
    for field in spec["fields"]:
        fn = field["name"]
        col = mapping.get(fn)
        rawval = raw.get(col) if col else None
        has_value = rawval is not None and str(rawval).strip() != ""
        if not has_value:
            if field.get("required"):
                errs.append(f"{field['label']} is required")
            elif "default" in field:
                out[fn] = field["default"]
            continue
        try:
            out[fn] = _coerce(rawval, field["type"])
        except (ValueError, TypeError):
            errs.append(f"{field['label']} has invalid value '{rawval}'")
    return out, errs


# --------------------------------------------------------------------------- #
# Commit
# --------------------------------------------------------------------------- #
def commit_import(
    db: Session,
    entity: str,
    rows: list[dict],
    mapping: dict[str, str | None],
    source_name: str,
    source_type: str,
    user_id: int | None = None,
) -> dict:
    spec = entity_spec(entity)
    started = _utcnow()
    imported = 0
    rejected = 0
    errors: list[dict] = []

    # caches for FK resolution
    wh_by_name = {w.name.lower(): w for w in db.scalars(select(Warehouse)).all()} if entity == "inventory" else {}
    prod_by_sku = {p.sku.lower(): p for p in db.scalars(select(Product)).all()} if entity == "inventory" else {}

    for idx, raw in enumerate(rows):
        norm, row_errors = _transform_row(spec, raw, mapping)
        if row_errors:
            rejected += 1
            if len(errors) < 50:
                errors.append({"row": idx + 1, "errors": row_errors})
            continue
        try:
            obj = _build_entity(entity, norm, wh_by_name, prod_by_sku)
            db.add(obj)
            imported += 1
        except Exception as exc:  # row-level failure
            rejected += 1
            if len(errors) < 50:
                errors.append({"row": idx + 1, "errors": [str(exc)]})

    try:
        db.commit()
    except Exception as exc:
        db.rollback()
        imported, rejected = 0, len(rows)
        errors = [{"row": 0, "errors": [f"Database commit failed: {exc}"]}]

    duration_ms = int((_utcnow() - started).total_seconds() * 1000)
    status = (
        ImportStatus.SUCCESS if rejected == 0 and imported > 0
        else ImportStatus.FAILED if imported == 0
        else ImportStatus.PARTIAL
    )
    job = ImportJob(
        source_name=source_name,
        source_type=source_type,
        entity_type=entity,
        status=status,
        rows_processed=len(rows),
        rows_imported=imported,
        rows_rejected=rejected,
        duration_ms=duration_ms,
        error_summary=errors or None,
        user_id=user_id,
    )
    db.add(job)
    db.commit()
    db.refresh(job)

    return {
        "job_id": job.id,
        "entity": entity,
        "status": status.value,
        "rows_processed": len(rows),
        "rows_imported": imported,
        "rows_rejected": rejected,
        "duration_ms": duration_ms,
        "errors": errors,
    }


def _build_entity(entity: str, norm: dict, wh_by_name: dict, prod_by_sku: dict):
    if entity == "suppliers":
        return Supplier(**norm)
    if entity == "warehouses":
        return Warehouse(**norm)
    if entity == "products":
        return Product(**norm)
    if entity == "shipments":
        status = norm.pop("status", "in_transit")
        try:
            status_enum = ShipmentStatus(status)
        except ValueError:
            status_enum = ShipmentStatus.IN_TRANSIT
        now = _utcnow()
        return Shipment(
            **norm,
            status=status_enum,
            current_location=norm.get("origin", ""),
            shipped_at=now,
            eta=now + timedelta(days=14),
        )
    if entity == "inventory":
        wh = wh_by_name.get(str(norm.pop("warehouse_name", "")).lower())
        prod = prod_by_sku.get(str(norm.pop("product_sku", "")).lower())
        if not wh:
            raise ValueError("warehouse not found")
        if not prod:
            raise ValueError("product SKU not found")
        return Inventory(warehouse_id=wh.id, product_id=prod.id, is_current=True, **norm)
    raise ValueError(f"Unsupported entity {entity}")


# --------------------------------------------------------------------------- #
# Operating mode
# --------------------------------------------------------------------------- #
def get_mode(db: Session) -> str:
    row = db.get(AppSetting, "operating_mode")
    return row.value if row else OperatingMode.DEMO.value


def set_mode(db: Session, mode: str) -> str:
    mode = OperatingMode(mode).value
    row = db.get(AppSetting, "operating_mode")
    if row:
        row.value = mode
    else:
        db.add(AppSetting(key="operating_mode", value=mode))
    db.commit()
    return mode


# --------------------------------------------------------------------------- #
# Connector sync simulation
# --------------------------------------------------------------------------- #
def sync_connector(db: Session, source: DataSource) -> dict:
    """Simulate a connector sync run and log it to the pipeline."""
    started = _utcnow()
    rng = random.Random(source.id * 7 + int(started.timestamp()))
    processed = rng.randint(80, 1200)
    failed = rng.randint(0, max(1, processed // 40))
    imported = processed - failed

    source.last_sync_at = started
    source.record_count = (source.record_count or 0) + imported
    source.status = ConnectorStatus.CONNECTED
    source.health = ConnectorHealth.HEALTHY if failed == 0 else ConnectorHealth.DEGRADED

    job = ImportJob(
        source_name=source.name,
        source_type="connector",
        entity_type="mixed",
        status=ImportStatus.SUCCESS if failed == 0 else ImportStatus.PARTIAL,
        rows_processed=processed,
        rows_imported=imported,
        rows_rejected=failed,
        duration_ms=rng.randint(400, 9000),
    )
    db.add(job)
    db.commit()
    db.refresh(source)
    return {
        "source": source.name,
        "records_processed": processed,
        "records_imported": imported,
        "records_failed": failed,
        "synced_at": started.isoformat(),
    }


# --------------------------------------------------------------------------- #
# Dashboard summary + AI context
# --------------------------------------------------------------------------- #
def summary(db: Session) -> dict:
    sources = db.scalars(select(DataSource)).all()
    connected = [s for s in sources if s.status == ConnectorStatus.CONNECTED]
    configured_types = {s.connector_type.value for s in sources}
    available = [t for t in INTEGRATION_TEMPLATES if t["type"] not in configured_types]

    last_sync = max((s.last_sync_at for s in sources if s.last_sync_at), default=None)
    total_records = sum(s.record_count for s in sources)

    today = _utcnow().date()
    imported_today = db.scalar(
        select(func.coalesce(func.sum(ImportJob.rows_imported), 0)).where(
            func.date(ImportJob.created_at) == today.isoformat()
        )
    ) or 0

    status_counts: dict[str, int] = {}
    for s in sources:
        status_counts[s.status.value] = status_counts.get(s.status.value, 0) + 1

    failures = [
        {"source": s.name, "health": s.health.value, "status": s.status.value}
        for s in sources
        if s.status == ConnectorStatus.ERROR or s.health == ConnectorHealth.DOWN
    ]

    return {
        "mode": get_mode(db),
        "connected_systems": len(connected),
        "total_sources": len(sources),
        "available_integrations": len(available),
        "last_sync_at": last_sync.isoformat() if last_sync else None,
        "records_imported_total": int(total_records),
        "records_imported_today": int(imported_today),
        "status_counts": status_counts,
        "failures": failures,
    }


def ai_context(db: Session) -> dict:
    """Compact data-source snapshot for the Operations Copilot."""
    s = summary(db)
    sources = db.scalars(select(DataSource)).all()
    return {
        "mode": s["mode"],
        "connected_systems": s["connected_systems"],
        "available_integrations": s["available_integrations"],
        "last_sync_at": s["last_sync_at"],
        "records_imported_today": s["records_imported_today"],
        "records_imported_total": s["records_imported_total"],
        "sources": [
            {
                "name": d.name,
                "type": d.connector_type.value,
                "status": d.status.value,
                "health": d.health.value,
                "last_sync_at": d.last_sync_at.isoformat() if d.last_sync_at else None,
                "record_count": d.record_count,
            }
            for d in sources
        ],
        "failures": s["failures"],
    }


# --------------------------------------------------------------------------- #
# Seeding mock connectors + pipeline history (idempotent)
# --------------------------------------------------------------------------- #
def seed_data_sources(db: Session) -> int:
    if db.scalar(select(func.count(DataSource.id))) or 0:
        return 0

    now = _utcnow()
    seeds = [
        ("SAP ERP — Production", ConnectorType.SAP_ERP, ConnectorStatus.CONNECTED, ConnectorHealth.HEALTHY,
         48230, now - timedelta(minutes=14), "https://sap.internal/api", "OAuth2", "Hourly"),
        ("Manhattan WMS", ConnectorType.WMS, ConnectorStatus.CONNECTED, ConnectorHealth.HEALTHY,
         26110, now - timedelta(minutes=42), "https://wms.internal/v2", "API Key", "Every 15 min"),
        ("project44 TMS", ConnectorType.TMS, ConnectorStatus.SYNCING, ConnectorHealth.HEALTHY,
         15890, now - timedelta(minutes=3), "https://api.project44.com", "API Key", "Real-time"),
        ("Salesforce CRM", ConnectorType.SALESFORCE_CRM, ConnectorStatus.ERROR, ConnectorHealth.DOWN,
         9120, now - timedelta(hours=6), "https://login.salesforce.com", "OAuth2", "Daily"),
    ]
    for name, ctype, status, health, rec, last, url, auth, freq in seeds:
        db.add(DataSource(
            name=name, connector_type=ctype, status=status, health=health,
            record_count=rec, last_sync_at=last, base_url=url, auth_method=auth,
            sync_frequency=freq, api_key_masked="••••••••" + str(random.randint(1000, 9999)),
            webhook_url=None, is_active=True,
        ))
    db.flush()

    # a little pipeline history so the monitor isn't empty
    rng = random.Random(42)
    history_sources = [
        ("SAP ERP — Production", "connector", "mixed"),
        ("Manhattan WMS", "connector", "inventory"),
        ("project44 TMS", "connector", "shipments"),
        ("Salesforce CRM", "connector", "mixed"),
    ]
    for i in range(14):
        name, stype, ent = rng.choice(history_sources)
        processed = rng.randint(60, 1500)
        failed = rng.randint(0, processed // 30)
        status = ImportStatus.SUCCESS if failed == 0 else (ImportStatus.PARTIAL if failed < processed else ImportStatus.FAILED)
        job = ImportJob(
            source_name=name, source_type=stype, entity_type=ent, status=status,
            rows_processed=processed, rows_imported=processed - failed, rows_rejected=failed,
            duration_ms=rng.randint(300, 12000),
        )
        job.created_at = now - timedelta(hours=i * 5 + rng.randint(0, 4))
        db.add(job)

    db.commit()
    return len(seeds)
